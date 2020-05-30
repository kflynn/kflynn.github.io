import sys

from openpyxl import load_workbook

def crunch_path(path: str) -> None:
    wb = load_workbook(path)
    sheet = None

    try:
        sheet = wb['Hospital COVID census']
    except KeyError:
        pass

    if not sheet:
        try:
            sheet = wb['Hospital COVID Census']
        except KeyError:
            pass

    if not sheet:
        raise Exception(f"{path} has no Hospital COVID Census sheet")

    for cell_id, wanted in [
        ( "A1", "Hospital Name" ),
        ( "C1", "Hospitalized Total COVID patients - suspected and confirmed (including ICU)" ),
        ( "D1", "Hospitalized COVID patients in ICU - suspected and confirmed" )
    ]:
        cell_name = sheet[cell_id].value.lower().replace('\n', ' ').replace('  ', ' ')
        wanted = wanted.lower()

        if cell_name != wanted:
            raise Exception(f"Cell {cell_id} should be '{wanted}' but is '{cell_name}'")

    have_zip = True

    cell_name = sheet["B1"].value.lower()

    if cell_name == "hospital county":
        have_zip = False
    elif cell_name != "hospital county and zip code":
        raise Exception(f"Cell B1 should be 'Hospital County' but is '{cell_name}'")

    counties = {}
    massachusetts = 0
    ma_other = 0

    for row in sheet[2:sheet.max_row]:
        name = row[0].value

        if not name:
            continue

        c_and_z = row[1].value
        confirmed = row[2].value
        icu = row[3].value

        # print(f"{name} ({c_and_z}): {confirmed}, {icu} ICU")

        county = c_and_z

        if have_zip:
            county, zip = c_and_z.split('-')
            county = county.strip()
            zip = zip.strip()

        if county not in counties:
            counties[county] = 0
        
        counties[county] += confirmed
        massachusetts += confirmed

        if (county != "Middlesex") and (county != "Suffolk"):
            ma_other += confirmed

    print("---- counties")

    for county in sorted(counties.keys()):
        print("%s: %d" % (county, counties[county]))

    print("---- total %d" % massachusetts)

    print("Massachusetts without Middlesex and Suffolk: %d" % ma_other)

if __name__ == "__main__":
    crunch_path(sys.argv[1])
